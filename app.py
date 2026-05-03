from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import tempfile
from pathlib import Path
from functools import wraps
import jwt
import bcrypt
import psycopg2
import psycopg2.extras

app = Flask(__name__, static_folder='../frontend', static_url_path='')
CORS(app)

DATABASE_URL = os.environ.get('DATABASE_URL')
FRONTEND_DIR = Path(__file__).resolve().parent.parent / 'frontend'
JWT_SECRET = os.environ.get('JWT_SECRET', 'chave-secreta-financeiro')
JWT_EXPIRATION_DAYS = 7

MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
         "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

# ── Banco de dados ────────────────────────────────────────────────────────────

def get_conn():
    """Retorna conexão com o PostgreSQL."""
    return psycopg2.connect(DATABASE_URL)

def init_db():
    """Cria as tabelas se não existirem."""
    conn = get_conn()
    cur  = conn.cursor()
    lock_id = 734512901

    try:
        # Evita corrida entre workers ao criar schema no startup.
        cur.execute("SELECT pg_advisory_lock(%s)", (lock_id,))

        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id         SERIAL PRIMARY KEY,
                nome       VARCHAR(100) NOT NULL,
                email      VARCHAR(150) UNIQUE NOT NULL,
                senha_hash VARCHAR(255) NOT NULL,
                criado_em  TIMESTAMP DEFAULT NOW()
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS lancamentos (
                id         SERIAL PRIMARY KEY,
                user_id    INTEGER REFERENCES usuarios(id) ON DELETE CASCADE,
                mes        VARCHAR(20),
                data       VARCHAR(20),
                descricao  TEXT NOT NULL,
                categoria  VARCHAR(50),
                tipo       VARCHAR(10) NOT NULL,
                valor      NUMERIC(12,2) NOT NULL,
                pagamento  VARCHAR(50),
                obs        TEXT,
                criado_em  TIMESTAMP DEFAULT NOW()
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS fixos (
                id               SERIAL PRIMARY KEY,
                user_id          INTEGER REFERENCES usuarios(id) ON DELETE CASCADE,
                tipo             VARCHAR(10) NOT NULL,
                descricao        TEXT NOT NULL,
                valor            NUMERIC(12,2) NOT NULL,
                categoria        VARCHAR(50),
                pagamento        VARCHAR(50),
                parcelado        BOOLEAN DEFAULT FALSE,
                parcela_atual    INTEGER DEFAULT 1,
                parcela_total    INTEGER DEFAULT 1,
                data_inicio      VARCHAR(10),
                ativo            BOOLEAN DEFAULT TRUE,
                criado_em        TIMESTAMP DEFAULT NOW()
            );
        """)

        # Migra schema de usuários para estruturas existentes.
        cur.execute("ALTER TABLE lancamentos ADD COLUMN IF NOT EXISTS user_id INTEGER")
        cur.execute("ALTER TABLE fixos ADD COLUMN IF NOT EXISTS user_id INTEGER")
        
        # Migra schema de lançamentos parcelados
        cur.execute("ALTER TABLE lancamentos ADD COLUMN IF NOT EXISTS parcelado BOOLEAN DEFAULT FALSE")
        cur.execute("ALTER TABLE lancamentos ADD COLUMN IF NOT EXISTS parcela_atual INTEGER")
        cur.execute("ALTER TABLE lancamentos ADD COLUMN IF NOT EXISTS total_parcelas INTEGER")
        cur.execute("ALTER TABLE lancamentos ADD COLUMN IF NOT EXISTS id_grupo_parcela VARCHAR(100)")

        # Migra schema antigo de fixos sem exigir recriacao de tabela.
        cur.execute("ALTER TABLE fixos ADD COLUMN IF NOT EXISTS parcelado BOOLEAN DEFAULT FALSE")
        cur.execute("ALTER TABLE fixos ADD COLUMN IF NOT EXISTS parcela_atual INTEGER DEFAULT 1")
        cur.execute("ALTER TABLE fixos ADD COLUMN IF NOT EXISTS parcela_total INTEGER DEFAULT 1")
        cur.execute("ALTER TABLE fixos ADD COLUMN IF NOT EXISTS data_inicio VARCHAR(10)")
        cur.execute("ALTER TABLE fixos ADD COLUMN IF NOT EXISTS ativo BOOLEAN DEFAULT TRUE")
        cur.execute("ALTER TABLE fixos ADD COLUMN IF NOT EXISTS criado_em TIMESTAMP DEFAULT NOW()")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS lancamentos_gerados (
                id               SERIAL PRIMARY KEY,
                fixo_id          INTEGER REFERENCES fixos(id) ON DELETE CASCADE,
                user_id          INTEGER REFERENCES usuarios(id) ON DELETE CASCADE,
                mes              VARCHAR(20),
                ano              INTEGER,
                valor            NUMERIC(12,2),
                gerado_em        TIMESTAMP DEFAULT NOW()
            );
        """)

        cur.execute("ALTER TABLE lancamentos_gerados ADD COLUMN IF NOT EXISTS user_id INTEGER")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_lancamentos_user_id ON lancamentos(user_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_fixos_user_id ON fixos(user_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_lancamentos_gerados_user_id ON lancamentos_gerados(user_id)")

        conn.commit()
    finally:
        try:
            cur.execute("SELECT pg_advisory_unlock(%s)", (lock_id,))
            conn.commit()
        except Exception:
            conn.rollback()
        cur.close()
        conn.close()


def migrar_meses():
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, data FROM lancamentos WHERE mes NOT LIKE '% %'")
        rows = cur.fetchall()
        for r in rows:
            if r[1]:
                novo_mes = mes_da_data(r[1])
                if novo_mes:
                    cur.execute("UPDATE lancamentos SET mes=%s WHERE id=%s", (novo_mes, r[0]))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print(f"  ❌ Erro ao migrar meses: {e}")

def bootstrap_database():
    """Inicializa o banco no startup (incluindo deploy com Gunicorn)."""
    if not DATABASE_URL:
        print("  ⚠️  DATABASE_URL não definida!")
        return

    try:
        init_db()
        migrar_meses()
        print("  ✅ Tabelas verificadas/criadas")
    except Exception as e:
        print(f"  ❌ Erro ao iniciar banco: {e}")

# Garante init do banco também quando o app roda via Gunicorn.
bootstrap_database()

# ── Helpers ───────────────────────────────────────────────────────────────────

def mes_da_data(data_str):
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            dt = datetime.strptime(data_str.strip(), fmt)
            return f"{MESES[dt.month - 1]} {dt.year}"
        except ValueError:
            continue
    return ""

def validar_email(email):
    return bool(re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', email or ''))

def validar_senha(senha):
    if len(senha or '') < 8:
        return False
    return (
        re.search(r'[A-Z]', senha)
        and re.search(r'[a-z]', senha)
        and re.search(r'\d', senha)
    )

def gerar_token(user_id, email):
    payload = {
        'user_id': user_id,
        'email': email,
        'exp': datetime.utcnow() + timedelta(days=JWT_EXPIRATION_DAYS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm='HS256')

def requer_token(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '').strip()
        if not token:
            return jsonify({'sucesso': False, 'mensagem': 'Token não fornecido'}), 401
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
            request.user_id = payload['user_id']
        except jwt.ExpiredSignatureError:
            return jsonify({'sucesso': False, 'mensagem': 'Sessão expirada. Faça login novamente.'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'sucesso': False, 'mensagem': 'Token inválido'}), 401
        return f(*args, **kwargs)
    return decorated

def verificar_e_gerar_parcelas(mes, ano, user_id):
    """Verifica fixos parcelados e gera lançamentos se necessário."""
    conn = get_conn()
    cur  = conn.cursor()
    
    # Busca todos os fixos parcelados ativos
    cur.execute("""
        SELECT id, valor, parcela_atual, parcela_total, data_inicio, tipo
        FROM fixos WHERE user_id=%s AND parcelado=TRUE AND ativo=TRUE
    """, (user_id,))
    fixos_parcelados = cur.fetchall()
    
    for fixo_id, valor, parcela_atual, parcela_total, data_inicio, tipo in fixos_parcelados:
        if not data_inicio:
            continue
        
        # Verifica se já foi gerado para este mês
        cur.execute("""
            SELECT id FROM lancamentos_gerados 
            WHERE fixo_id=%s AND user_id=%s AND mes=%s AND ano=%s
        """, (fixo_id, user_id, mes, ano))
        
        if cur.fetchone():
            # Já foi gerado, pula
            continue
        
        # Calcula se este mês está no range de parcelas
        # data_inicio é "01/05/2026" (1º dia do mês inicial)
        try:
            dt_inicio = datetime.strptime(data_inicio, "%d/%m/%Y")
            mes_inicio = dt_inicio.month
            ano_inicio = dt_inicio.year
            
            # Calcula quantos meses passaram desde o início
            meses_decorridos = (ano - ano_inicio) * 12 + (MESES.index(mes) - (mes_inicio - 1))
            
            # Se dentro do range de parcelas
            if 0 <= meses_decorridos < parcela_total:
                # Gera lançamento
                cur.execute("""
                    INSERT INTO lancamentos_gerados (fixo_id, user_id, mes, ano, valor)
                    VALUES (%s, %s, %s, %s, %s)
                """, (fixo_id, user_id, mes, ano, valor))
                
                # Atualiza parcela_atual
                nova_parcela = meses_decorridos + 1
                cur.execute("""
                    UPDATE fixos SET parcela_atual=%s WHERE id=%s AND user_id=%s
                """, (nova_parcela, fixo_id, user_id))
                
                # Se chegou no final, desativa
                if nova_parcela >= parcela_total:
                    cur.execute("UPDATE fixos SET ativo=FALSE WHERE id=%s AND user_id=%s", (fixo_id, user_id))
        except ValueError:
            continue
    
    conn.commit(); cur.close(); conn.close()

def totais_fixos(user_id):
    """Retorna total de rendas e gastos fixos (apenas recorrentes, sem parcelados)."""
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("SELECT tipo, SUM(valor) FROM fixos WHERE user_id=%s AND ativo=TRUE AND parcelado=FALSE GROUP BY tipo", (user_id,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    resultado = {'ENTRADA': 0.0, 'SAÍDA': 0.0}
    for tipo, total in rows:
        resultado[tipo] = float(total or 0)
    return resultado

# ── Rotas API ─────────────────────────────────────────────────────────────────

@app.route('/api/health', methods=['GET'])
def healthcheck():
    """Healthcheck simples para validar API e conexão com banco."""
    db_ok = False
    db_error = None

    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute('SELECT 1')
        cur.fetchone()
        cur.close()
        conn.close()
        db_ok = True
    except Exception as e:
        db_error = str(e)

    status = 200 if db_ok else 503
    return jsonify({
        'api': 'ok',
        'database': 'ok' if db_ok else 'error',
        'database_error': db_error
    }), status

@app.route('/api/auth/cadastro', methods=['POST'])
def cadastro():
    try:
        dados = request.json or {}
        nome = str(dados.get('nome', '')).strip()
        email = str(dados.get('email', '')).strip().lower()
        senha = str(dados.get('senha', ''))

        if len(nome) < 2:
            return jsonify({'sucesso': False, 'mensagem': 'Nome deve ter pelo menos 2 caracteres.'}), 400
        if not validar_email(email):
            return jsonify({'sucesso': False, 'mensagem': 'E-mail inválido.'}), 400
        if not validar_senha(senha):
            return jsonify({'sucesso': False, 'mensagem': 'Senha deve ter no mínimo 8 caracteres, com maiúscula, minúscula e número.'}), 400

        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cur.execute('SELECT id FROM usuarios WHERE LOWER(email)=LOWER(%s)', (email,))
        if cur.fetchone():
            cur.close(); conn.close()
            return jsonify({'sucesso': False, 'mensagem': 'Este e-mail já está cadastrado.'}), 409

        senha_hash = bcrypt.hashpw(senha.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        cur.execute(
            '''
            INSERT INTO usuarios (nome, email, senha_hash)
            VALUES (%s, %s, %s)
            RETURNING id, nome, email
            ''',
            (nome, email, senha_hash)
        )
        usuario = cur.fetchone()
        conn.commit()
        cur.close(); conn.close()

        token = gerar_token(usuario['id'], usuario['email'])
        return jsonify({
            'sucesso': True,
            'mensagem': 'Cadastro realizado com sucesso!',
            'token': token,
            'usuario': {
                'id': usuario['id'],
                'nome': usuario['nome'],
                'email': usuario['email']
            }
        }), 201

    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        dados = request.json or {}
        email = str(dados.get('email', '')).strip().lower()
        senha = str(dados.get('senha', ''))

        if not email or not senha:
            return jsonify({'sucesso': False, 'mensagem': 'E-mail e senha são obrigatórios.'}), 400

        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            '''
            SELECT id, nome, email, senha_hash
            FROM usuarios
            WHERE LOWER(email)=LOWER(%s)
            ''',
            (email,)
        )
        usuario = cur.fetchone()
        cur.close(); conn.close()

        credenciais_invalidas = (
            not usuario
            or not bcrypt.checkpw(senha.encode('utf-8'), usuario['senha_hash'].encode('utf-8'))
        )
        if credenciais_invalidas:
            return jsonify({'sucesso': False, 'mensagem': 'E-mail ou senha incorretos'}), 401

        token = gerar_token(usuario['id'], usuario['email'])
        return jsonify({
            'sucesso': True,
            'mensagem': 'Login realizado com sucesso!',
            'token': token,
            'usuario': {
                'id': usuario['id'],
                'nome': usuario['nome'],
                'email': usuario['email']
            }
        }), 200
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500

@app.route('/api/adicionar-lancamento', methods=['POST'])
@requer_token
def adicionar_lancamento():
    try:
        dados_input = request.json or {}
        lista_dados = dados_input if isinstance(dados_input, list) else [dados_input]
        
        conn = get_conn()
        cur  = conn.cursor()
        ids_inseridos = []
        mes_retorno = None

        for dados in lista_dados:
            for campo in ['data', 'descricao', 'categoria', 'tipo', 'valor']:
                if not str(dados.get(campo, '')).strip():
                    conn.rollback(); cur.close(); conn.close()
                    return jsonify({"erro": f"Campo obrigatório: {campo}"}), 400

            mes  = mes_da_data(str(dados['data']))
            if not mes_retorno: mes_retorno = mes
            tipo = dados['tipo'].upper()
            
            parcelado = bool(dados.get('parcelado', False))
            parcela_atual = int(dados.get('parcelaAtual')) if parcelado and dados.get('parcelaAtual') else None
            total_parcelas = int(dados.get('totalParcelas')) if parcelado and dados.get('totalParcelas') else None
            id_grupo_parcela = str(dados.get('idGrupoParcela', '')) if parcelado else None

            cur.execute("""
                INSERT INTO lancamentos (user_id, mes, data, descricao, categoria, tipo, valor, pagamento, obs, parcelado, parcela_atual, total_parcelas, id_grupo_parcela)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                request.user_id,
                mes, dados['data'], dados['descricao'], dados['categoria'],
                tipo, float(dados['valor']),
                dados.get('pagamento', ''), dados.get('obs', ''),
                parcelado, parcela_atual, total_parcelas, id_grupo_parcela
            ))
            ids_inseridos.append(cur.fetchone()[0])
            
        conn.commit(); cur.close(); conn.close()

        return jsonify({"sucesso": True, "mensagem": "Lançamento(s) adicionado(s)!", "id": ids_inseridos[0] if ids_inseridos else None, "mes": mes_retorno}), 201

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/lancamentos', methods=['GET'])
@requer_token
def listar_lancamentos():
    try:
        mes_filtro = request.args.get('mes', '').strip()
        conn = get_conn()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        if mes_filtro:
            cur.execute("""
                SELECT * FROM lancamentos WHERE user_id=%s AND LOWER(mes)=LOWER(%s)
                ORDER BY criado_em DESC
            """, (request.user_id, mes_filtro))
        else:
            cur.execute("SELECT * FROM lancamentos WHERE user_id=%s ORDER BY criado_em DESC", (request.user_id,))

        rows = cur.fetchall()
        cur.close(); conn.close()

        lancamentos = []
        for r in rows:
            lancamentos.append({
                "id":               r['id'],
                "mes":              r['mes'],
                "data":             r['data'],
                "descricao":        r['descricao'],
                "categoria":        r['categoria'],
                "tipo":             r['tipo'],
                "valor":            float(r['valor']),
                "pagamento":        r['pagamento'],
                "obs":              r['obs'],
                "parcelado":        r.get('parcelado', False),
                "parcela_atual":    r.get('parcela_atual', None),
                "total_parcelas":   r.get('total_parcelas', None),
                "id_grupo_parcela": r.get('id_grupo_parcela', None)
            })

        def sort_key(x):
            date_str = x['data'] or ''
            dt = None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
                try:
                    dt = datetime.strptime(date_str.strip(), fmt)
                    break
                except (ValueError, AttributeError):
                    pass
            return (dt or datetime.min, x['id'])

        lancamentos.sort(key=sort_key, reverse=True)

        return jsonify({"lancamentos": lancamentos, "total": len(lancamentos)}), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/lancamentos/<int:lancamento_id>', methods=['PUT'])
@requer_token
def editar_lancamento(lancamento_id):
    try:
        dados = request.json or {}

        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM lancamentos WHERE id=%s AND user_id=%s", (lancamento_id, request.user_id))
        atual = cur.fetchone()

        if not atual:
            cur.close(); conn.close()
            return jsonify({"erro": "Lançamento não encontrado"}), 404

        data = str(dados.get('data', atual['data'])).strip()
        descricao = str(dados.get('descricao', atual['descricao'])).strip()
        categoria = str(dados.get('categoria', atual['categoria'] or '')).strip()
        tipo = str(dados.get('tipo', atual['tipo'])).upper().strip()
        pagamento = str(dados.get('pagamento', atual['pagamento'] or '')).strip()
        obs = str(dados.get('obs', atual['obs'] or '')).strip()

        valor_recebido = dados.get('valor', atual['valor'])
        try:
            valor = float(valor_recebido)
        except (TypeError, ValueError):
            cur.close(); conn.close()
            return jsonify({"erro": "Valor inválido"}), 400

        if not data or not descricao or not categoria or tipo not in ('ENTRADA', 'SAÍDA') or valor <= 0:
            cur.close(); conn.close()
            return jsonify({"erro": "Dados inválidos para atualização"}), 400

        mes = mes_da_data(data)
        if not mes:
            cur.close(); conn.close()
            return jsonify({"erro": "Data inválida"}), 400

        abrangencia = str(dados.get('abrangencia', 'UNICA')).upper()
        
        cur2 = conn.cursor()
        
        if atual['parcelado'] and atual['id_grupo_parcela'] and abrangencia != 'UNICA':
            cur2.execute("""
                UPDATE lancamentos
                SET mes=%s, data=%s, descricao=%s, categoria=%s, tipo=%s, valor=%s, pagamento=%s, obs=%s
                WHERE id=%s AND user_id=%s
            """, (mes, data, descricao, categoria, tipo, valor, pagamento, obs, lancamento_id, request.user_id))
            
            if abrangencia == 'TODAS':
                cur2.execute("""
                    UPDATE lancamentos
                    SET descricao=%s, categoria=%s, tipo=%s, valor=%s, pagamento=%s, obs=%s
                    WHERE id_grupo_parcela=%s AND user_id=%s AND id != %s
                """, (descricao, categoria, tipo, valor, pagamento, obs, atual['id_grupo_parcela'], request.user_id, lancamento_id))
            elif abrangencia == 'PROXIMAS':
                cur2.execute("""
                    UPDATE lancamentos
                    SET descricao=%s, categoria=%s, tipo=%s, valor=%s, pagamento=%s, obs=%s
                    WHERE id_grupo_parcela=%s AND user_id=%s AND parcela_atual > %s
                """, (descricao, categoria, tipo, valor, pagamento, obs, atual['id_grupo_parcela'], request.user_id, atual['parcela_atual']))
        else:
            cur2.execute("""
                UPDATE lancamentos
                SET mes=%s, data=%s, descricao=%s, categoria=%s, tipo=%s, valor=%s, pagamento=%s, obs=%s
                WHERE id=%s AND user_id=%s
            """, (mes, data, descricao, categoria, tipo, valor, pagamento, obs, lancamento_id, request.user_id))

        conn.commit()
        cur2.close(); cur.close(); conn.close()

        return jsonify({"sucesso": True, "mes": mes}), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/lancamentos/<int:lancamento_id>', methods=['DELETE'])
@requer_token
def deletar_lancamento(lancamento_id):
    try:
        abrangencia = request.args.get('abrangencia', 'UNICA').upper()
        conn = get_conn()
        cur = conn.cursor()
        
        cur.execute("SELECT parcelado, id_grupo_parcela, parcela_atual FROM lancamentos WHERE id=%s AND user_id=%s", (lancamento_id, request.user_id))
        atual = cur.fetchone()
        
        if not atual:
            conn.rollback(); cur.close(); conn.close()
            return jsonify({"erro": "Lançamento não encontrado"}), 404
            
        if atual[0] and atual[1] and abrangencia != 'UNICA':
            if abrangencia == 'TODAS':
                cur.execute("DELETE FROM lancamentos WHERE id_grupo_parcela=%s AND user_id=%s", (atual[1], request.user_id))
            elif abrangencia == 'PROXIMAS':
                cur.execute("DELETE FROM lancamentos WHERE id_grupo_parcela=%s AND user_id=%s AND parcela_atual >= %s", (atual[1], request.user_id, atual[2]))
        else:
            cur.execute("DELETE FROM lancamentos WHERE id=%s AND user_id=%s", (lancamento_id, request.user_id))

        conn.commit()
        cur.close(); conn.close()
        return jsonify({"sucesso": True}), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/resumo', methods=['GET'])
@requer_token
def resumo():
    try:
        mes_filtro = request.args.get('mes', '').strip()
        ano_filtro = request.args.get('ano', str(datetime.now().year)).strip()
        
        conn = get_conn()
        cur  = conn.cursor()

        # Gera parcelas se necessário
        if mes_filtro:
            partes = mes_filtro.split()
            nome_mes = partes[0]
            ano = int(partes[1]) if len(partes) > 1 else int(ano_filtro) if ano_filtro else datetime.now().year
            verificar_e_gerar_parcelas(nome_mes, ano, request.user_id)

        # Conta quantos meses existem para média
        cur.execute("SELECT COUNT(DISTINCT mes) FROM lancamentos WHERE user_id=%s", (request.user_id,))
        num_meses = cur.fetchone()[0] or 1

        if mes_filtro:
            cur.execute("""
                SELECT tipo, SUM(valor) FROM lancamentos
                WHERE user_id=%s AND LOWER(mes)=LOWER(%s) GROUP BY tipo
            """, (request.user_id, mes_filtro))
        else:
            cur.execute("SELECT tipo, SUM(valor) FROM lancamentos WHERE user_id=%s GROUP BY tipo", (request.user_id,))

        rows = cur.fetchall()
        
        entradas = saidas = 0.0
        for tipo, total in rows:
            if tipo == 'ENTRADA': entradas = float(total or 0)
            elif tipo == 'SAÍDA': saidas   = float(total or 0)

        # Se for "Todos os meses", tiramos a média dos lançamentos
        if not mes_filtro:
            entradas = entradas / num_meses
            saidas = saidas / num_meses

        entrada_parc_val = 0.0
        saida_parc_val = 0.0

        # Adiciona lançamentos gerados de fixos parcelados
        if mes_filtro:
            cur.execute("""
                SELECT SUM(CASE WHEN f.tipo='ENTRADA' THEN lg.valor ELSE 0 END),
                       SUM(CASE WHEN f.tipo='SAÍDA' THEN lg.valor ELSE 0 END)
                FROM lancamentos_gerados lg
                JOIN fixos f ON lg.fixo_id = f.id
                WHERE lg.user_id=%s AND LOWER(lg.mes)=LOWER(%s) AND lg.ano=%s
            """, (request.user_id, nome_mes, ano))
            entrada_parc, saida_parc = cur.fetchone()
            if entrada_parc: entrada_parc_val = float(entrada_parc)
            if saida_parc: saida_parc_val = float(saida_parc)
            entradas += entrada_parc_val
            saidas += saida_parc_val
        else:
            # Para "Todos os meses", somamos todos os gerados e dividimos pelos meses
            cur.execute("""
                SELECT SUM(CASE WHEN f.tipo='ENTRADA' THEN lg.valor ELSE 0 END),
                       SUM(CASE WHEN f.tipo='SAÍDA' THEN lg.valor ELSE 0 END)
                FROM lancamentos_gerados lg
                JOIN fixos f ON lg.fixo_id = f.id
                WHERE lg.user_id=%s
            """, (request.user_id,))
            entrada_parc, saida_parc = cur.fetchone()
            if entrada_parc: entrada_parc_val = float(entrada_parc) / num_meses
            if saida_parc: saida_parc_val = float(saida_parc) / num_meses
            entradas += entrada_parc_val
            saidas += saida_parc_val

        fixos = totais_fixos(request.user_id)
        # Fixos já são mensais por natureza
        entradas_dashboard = entradas + fixos['ENTRADA']
        saidas_dashboard = saidas + fixos['SAÍDA']
        saldo_dashboard = entradas_dashboard - saidas_dashboard
        
        cur.close(); conn.close()

        return jsonify({
            "entradas":     round(entradas_dashboard, 2),
            "saidas":       round(saidas_dashboard, 2),
            "saldo":        round(saldo_dashboard, 2),
            "renda_fixa":   round(fixos['ENTRADA'] + entrada_parc_val, 2),
            "gastos_fixos": round(fixos['SAÍDA'] + saida_parc_val,   2),
            "num_meses":    num_meses
        }), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/categorias-resumo', methods=['GET'])
@requer_token
def categorias_resumo():
    try:
        mes_filtro = request.args.get('mes', '').strip()
        conn = get_conn()
        cur  = conn.cursor()

        # Conta meses para média
        cur.execute("SELECT COUNT(DISTINCT mes) FROM lancamentos WHERE user_id=%s", (request.user_id,))
        num_meses = cur.fetchone()[0] or 1
        
        nome_mes = ""
        ano = datetime.now().year
        if mes_filtro:
            partes = mes_filtro.split()
            nome_mes = partes[0]
            if len(partes) > 1:
                ano = int(partes[1])

        # 1. Busca gastos de lançamentos manuais (sempre variáveis)
        if mes_filtro:
            cur.execute("""
                SELECT categoria, SUM(valor) FROM lancamentos
                WHERE user_id=%s AND tipo='SAÍDA' AND LOWER(mes)=LOWER(%s)
                GROUP BY categoria
            """, (request.user_id, mes_filtro))
        else:
            cur.execute("""
                SELECT categoria, SUM(valor) FROM lancamentos
                WHERE user_id=%s AND tipo='SAÍDA'
                GROUP BY categoria
            """, (request.user_id,))
        
        dict_variaveis = {r[0]: float(r[1]) for r in cur.fetchall()}

        # Se for "Todos os meses", tiramos a média dos lançamentos por categoria
        if not mes_filtro:
            for cat in dict_variaveis:
                dict_variaveis[cat] = dict_variaveis[cat] / num_meses

        dict_fixos = {}

        # 2. Busca gastos de fixos recorrentes e gerados (fixos)
        if mes_filtro:
            # Fixos recorrentes ativos
            cur.execute("""
                SELECT categoria, SUM(valor) FROM fixos
                WHERE user_id=%s AND tipo='SAÍDA' AND ativo=TRUE AND parcelado=FALSE
                GROUP BY categoria
            """, (request.user_id,))
            for cat, total in cur.fetchall():
                dict_fixos[cat] = dict_fixos.get(cat, 0.0) + float(total)

            # Gastos gerados (parcelados) para aquele mês
            cur.execute("""
                SELECT f.categoria, SUM(lg.valor)
                FROM lancamentos_gerados lg
                JOIN fixos f ON lg.fixo_id = f.id
                WHERE lg.user_id=%s AND f.tipo='SAÍDA' AND LOWER(lg.mes)=LOWER(%s) AND lg.ano=%s
                GROUP BY f.categoria
            """, (request.user_id, nome_mes, ano))
            for cat, total in cur.fetchall():
                dict_fixos[cat] = dict_fixos.get(cat, 0.0) + float(total)
        else:
            # No modo "Todos", somamos os fixos recorrentes ativos
            cur.execute("""
                SELECT categoria, SUM(valor) FROM fixos
                WHERE user_id=%s AND tipo='SAÍDA' AND ativo=TRUE AND parcelado=FALSE
                GROUP BY categoria
            """, (request.user_id,))
            for cat, total in cur.fetchall():
                dict_fixos[cat] = dict_fixos.get(cat, 0.0) + float(total)

            # E somamos a média dos parcelados já gerados
            cur.execute("""
                SELECT f.categoria, SUM(lg.valor)
                FROM lancamentos_gerados lg
                JOIN fixos f ON lg.fixo_id = f.id
                WHERE lg.user_id=%s AND f.tipo='SAÍDA'
                GROUP BY f.categoria
            """, (request.user_id,))
            for cat, total in cur.fetchall():
                dict_fixos[cat] = dict_fixos.get(cat, 0.0) + (float(total) / num_meses)

        cur.close(); conn.close()

        # Une os dicionários
        todas_categorias = set(list(dict_variaveis.keys()) + list(dict_fixos.keys()))
        lista_final = []
        for cat in todas_categorias:
            v = dict_variaveis.get(cat, 0.0)
            f = dict_fixos.get(cat, 0.0)
            lista_final.append({
                "categoria": cat,
                "total": v + f,
                "variavel": v,
                "fixo": f
            })

        lista_final.sort(key=lambda x: x['total'], reverse=True)
        return jsonify({"categorias": lista_final}), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/meses-disponiveis', methods=['GET'])
@requer_token
def meses_disponiveis():
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("SELECT DISTINCT mes FROM lancamentos WHERE user_id=%s AND mes IS NOT NULL", (request.user_id,))
        rows = cur.fetchall()
        meses_set = set(r[0] for r in rows if r[0])
        
        cur.execute("SELECT DISTINCT mes, ano FROM lancamentos_gerados WHERE user_id=%s AND mes IS NOT NULL", (request.user_id,))
        rows_lg = cur.fetchall()
        for r in rows_lg:
            if r[0] and r[1]:
                meses_set.add(f"{r[0]} {r[1]}")
                
        cur.close(); conn.close()

        def mes_key(m_str):
            partes = m_str.split()
            if len(partes) >= 2:
                try:
                    ano = int(partes[1])
                    mes_idx = MESES.index(partes[0].capitalize())
                    return (ano, mes_idx)
                except ValueError:
                    pass
            return (9999, 99)

        meses = sorted(list(meses_set), key=mes_key)
        return jsonify({"meses": meses}), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/listar-categorias', methods=['GET'])
@requer_token
def listar_categorias():
    return jsonify({
        "entrada": ["Salário","Freelance","Investimentos","Prêmio/Bônus","Outros (Entrada)"],
        "saida":   ["Alimentação","Fast Food","Mercado","Moradia","Transporte","Veiculo","Gasolina","Saúde","Educação",
                    "Lazer","Vestuário","Assinaturas","Beleza","Pet","Presentes","Contas/Serviços","Outros (Saída)"]
    }), 200


@app.route('/api/listar-pagamentos', methods=['GET'])
@requer_token
def listar_pagamentos():
    return jsonify({"pagamentos": [
        "Dinheiro","Cartão Débito","Cartão Crédito",
        "Transferência","Pix","Boleto","Débito Automático"
    ]}), 200


@app.route('/api/fixos', methods=['GET'])
@requer_token
def listar_fixos():
    try:
        conn = get_conn()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, tipo, descricao, valor, categoria, pagamento, 
                   parcelado, parcela_atual, parcela_total, data_inicio, ativo
            FROM fixos WHERE user_id=%s ORDER BY tipo, id
        """, (request.user_id,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        
        entradas = []
        saidas = []
        for r in rows:
            item = {
                "id": r['id'],
                "tipo": r['tipo'],
                "descricao": r['descricao'],
                "valor": float(r['valor']),
                "categoria": r['categoria'],
                "pagamento": r['pagamento'],
                "parcelado": r['parcelado'],
                "parcela_atual": r['parcela_atual'],
                "parcela_total": r['parcela_total'],
                "data_inicio": r['data_inicio'],
                "ativo": r['ativo']
            }
            if r['tipo'] == 'ENTRADA':
                entradas.append(item)
            else:
                saidas.append(item)
        
        return jsonify({"entradas": entradas, "saidas": saidas}), 200
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/fixos', methods=['POST'])
@requer_token
def adicionar_fixo():
    try:
        dados = request.json or {}
        
        # Validações
        for campo in ['tipo', 'descricao', 'valor', 'categoria']:
            if not str(dados.get(campo, '')).strip():
                return jsonify({"erro": f"Campo obrigatório: {campo}"}), 400
        
        tipo = dados['tipo'].upper()
        if tipo not in ['ENTRADA', 'SAÍDA']:
            return jsonify({"erro": "Tipo deve ser ENTRADA ou SAÍDA"}), 400
        
        parcelado = dados.get('parcelado', False)
        if parcelado:
            parcela_total = dados.get('parcela_total')
            data_inicio = dados.get('data_inicio', '')
            
            if not parcela_total or parcela_total < 2 or parcela_total > 120:
                return jsonify({"erro": "Parcelas devem estar entre 2 e 120"}), 400
            if not data_inicio or not data_inicio.strip():
                return jsonify({"erro": "Data de início é obrigatória para parcelados"}), 400
        
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO fixos (user_id, tipo, descricao, valor, categoria, pagamento, 
                              parcelado, parcela_atual, parcela_total, data_inicio, ativo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE) RETURNING id
        """, (
            request.user_id,
            tipo, dados['descricao'].strip(),
            float(dados['valor']), dados.get('categoria',''),
            dados.get('pagamento',''), parcelado,
            1, dados.get('parcela_total', 1), dados.get('data_inicio', '')
        ))
        novo_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return jsonify({"sucesso": True, "id": novo_id}), 201
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/fixos/<int:fixo_id>', methods=['PUT'])
@requer_token
def editar_fixo(fixo_id):
    try:
        dados = request.json or {}
        
        conn = get_conn()
        cur  = conn.cursor()
        
        # Busca o fixo atual
        cur.execute("SELECT * FROM fixos WHERE id=%s AND user_id=%s", (fixo_id, request.user_id))
        fixo = cur.fetchone()
        if not fixo:
            cur.close(); conn.close()
            return jsonify({"erro": "Fixo não encontrado"}), 404
        
        # Atualiza campos fornecidos
        campos = []
        valores = []
        
        if 'descricao' in dados:
            campos.append("descricao=%s")
            valores.append(dados['descricao'])
        if 'valor' in dados:
            campos.append("valor=%s")
            valores.append(float(dados['valor']))
        if 'categoria' in dados:
            campos.append("categoria=%s")
            valores.append(dados['categoria'])
        if 'pagamento' in dados:
            campos.append("pagamento=%s")
            valores.append(dados['pagamento'])
        if 'parcelado' in dados:
            campos.append("parcelado=%s")
            valores.append(dados['parcelado'])
        if 'parcela_total' in dados:
            campos.append("parcela_total=%s")
            valores.append(dados['parcela_total'])
        if 'data_inicio' in dados:
            campos.append("data_inicio=%s")
            valores.append(dados['data_inicio'])
        
        if campos:
            valores.append(fixo_id)
            valores.append(request.user_id)
            query = f"UPDATE fixos SET {', '.join(campos)} WHERE id=%s AND user_id=%s"
            cur.execute(query, valores)
            conn.commit()
        
        cur.close(); conn.close()
        return jsonify({"sucesso": True}), 200
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/fixos/<int:fixo_id>', methods=['DELETE'])
@requer_token
def deletar_fixo(fixo_id):
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("DELETE FROM fixos WHERE id=%s AND user_id=%s", (fixo_id, request.user_id))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"sucesso": True}), 200
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/api/fixos/<int:fixo_id>/toggle', methods=['PATCH'])
@requer_token
def toggle_fixo(fixo_id):
    try:
        conn = get_conn()
        cur  = conn.cursor()
        
        cur.execute("SELECT ativo FROM fixos WHERE id=%s AND user_id=%s", (fixo_id, request.user_id))
        result = cur.fetchone()
        if not result:
            cur.close(); conn.close()
            return jsonify({"erro": "Fixo não encontrado"}), 404
        
        novo_ativo = not result[0]
        cur.execute("UPDATE fixos SET ativo=%s WHERE id=%s AND user_id=%s", (novo_ativo, fixo_id, request.user_id))
        conn.commit(); cur.close(); conn.close()
        
        return jsonify({"sucesso": True, "ativo": novo_ativo}), 200
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# ── Exportar Excel ────────────────────────────────────────────────────────────

@app.route('/api/exportar-excel', methods=['GET'])
@requer_token
def exportar_excel():
    try:
        mes_filtro = request.args.get('mes', '').strip()

        # Busca dados do banco
        conn = get_conn()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        if mes_filtro:
            cur.execute("""
                SELECT * FROM lancamentos WHERE user_id=%s AND LOWER(mes)=LOWER(%s)
                ORDER BY criado_em ASC
            """, (request.user_id, mes_filtro))
        else:
            cur.execute("SELECT * FROM lancamentos WHERE user_id=%s ORDER BY criado_em ASC", (request.user_id,))

        lancamentos = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()

        fixos = totais_fixos(request.user_id)
        renda_fixa   = fixos['ENTRADA']
        gastos_fixos = fixos['SAÍDA']

        total_entrada = sum(float(l['valor']) for l in lancamentos if l['tipo'] == 'ENTRADA')
        total_saida   = sum(float(l['valor']) for l in lancamentos if l['tipo'] == 'SAÍDA')
        if mes_filtro:
            total_entrada += renda_fixa
            total_saida   += gastos_fixos
        saldo = total_entrada - total_saida

        # Estilos
        def bdr():
            s = Side(style='thin', color='CBD5E1')
            return Border(top=s, bottom=s, left=s, right=s)
        def lfill(c): return PatternFill('solid', fgColor=c)
        def ca():     return Alignment(horizontal='center', vertical='center', wrap_text=True)
        def la():     return Alignment(horizontal='left',   vertical='center')
        def ra():     return Alignment(horizontal='right',  vertical='center')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Lançamentos'

        for col, w in {'A':4,'B':14,'C':14,'D':32,'E':20,'F':12,'G':16,'H':22,'I':18}.items():
            ws.column_dimensions[col].width = w

        # Título
        titulo = f'CONTROLE FINANCEIRO — {mes_filtro.upper() if mes_filtro else "TODOS OS MESES"}'
        ws.row_dimensions[1].height = 44
        ws.merge_cells('A1:I1')
        c = ws['A1']; c.value = titulo
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=14)
        c.fill = lfill('1E293B'); c.alignment = ca()

        # Cards resumo
        ws.row_dimensions[2].height = 10
        ws.row_dimensions[3].height = 30
        for col_r, label, valor, bg, fc in [
            ('A3:C3','💚 ENTRADAS', total_entrada,'D1FAE5','065F46'),
            ('D3:F3','🔴 SAÍDAS',   total_saida,  'FEE2E2','7F1D1D'),
            ('G3:I3','💙 SALDO',    saldo,        'DBEAFE','1E3A8A'),
        ]:
            ws.merge_cells(col_r)
            c = ws[col_r.split(':')[0]]
            c.value = f'{label}:  R$ {valor:,.2f}'.replace(',','X').replace('.',',').replace('X','.')
            c.font = Font(name='Arial', bold=True, color=fc, size=11)
            c.fill = lfill(bg); c.alignment = ca(); c.border = bdr()

        ws.row_dimensions[4].height = 10

        # Cabeçalhos
        ws.row_dimensions[5].height = 28
        for ci, h in enumerate(['#','Mês','Data','Descrição','Categoria','Tipo','Valor (R$)','Forma Pagto.','Obs.'], 1):
            c = ws.cell(row=5, column=ci, value=h)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill = lfill('1E293B'); c.alignment = ca(); c.border = bdr()

        # Linhas
        for ri, l in enumerate(lancamentos, 6):
            ws.row_dimensions[ri].height = 22
            rf  = lfill('F8FAFC') if ri % 2 == 0 else lfill('E2E8F0')
            ent = l['tipo'].upper() == 'ENTRADA'
            ct  = '065F46' if ent else '7F1D1D'
            for ci, v in enumerate([
                ri-5, l['mes'], l['data'], l['descricao'],
                l['categoria'], l['tipo'], float(l['valor']),
                l['pagamento'], l['obs']
            ], 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = rf; c.border = bdr()
                if ci == 6:
                    c.font = Font(name='Arial', bold=True, color=ct, size=10); c.alignment = ca()
                elif ci == 7:
                    c.font = Font(name='Arial', bold=True, color=ct, size=10)
                    c.number_format = 'R$ #,##0.00'; c.alignment = ra()
                elif ci == 1:
                    c.font = Font(name='Arial', bold=True, size=10); c.alignment = ca()
                else:
                    c.font = Font(name='Arial', size=10)
                    c.alignment = la() if ci in (4,5,8,9) else ca()

        # Aba fixos
        ws2 = wb.create_sheet('Fixos Mensais')
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 20
        ws2.row_dimensions[1].height = 40
        ws2.merge_cells('A1:B1')
        c = ws2['A1']; c.value = 'RENDAS E GASTOS FIXOS MENSAIS'
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=13)
        c.fill = lfill('1E293B'); c.alignment = ca()
        for ri, (label, valor, bg, fc) in enumerate([
            ('💚 Renda Fixa',   renda_fixa,              'D1FAE5','065F46'),
            ('🔴 Gastos Fixos', gastos_fixos,            'FEE2E2','7F1D1D'),
            ('💙 Saldo Fixo',   renda_fixa-gastos_fixos, 'DBEAFE','1E3A8A'),
        ], 3):
            ws2.row_dimensions[ri].height = 28
            c = ws2.cell(row=ri, column=1, value=label)
            c.font = Font(name='Arial', bold=True, color=fc, size=11)
            c.fill = lfill(bg); c.alignment = la(); c.border = bdr()
            c = ws2.cell(row=ri, column=2, value=valor)
            c.number_format = 'R$ #,##0.00'
            c.font = Font(name='Arial', bold=True, color=fc, size=12)
            c.fill = lfill(bg); c.alignment = ra(); c.border = bdr()

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp.name); tmp.close()

        data_hoje = datetime.now().strftime('%d-%m-%Y')
        nome = f'controle_{mes_filtro or "completo"}_{data_hoje}.xlsx'

        return send_file(tmp.name, as_attachment=True, download_name=nome,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# ── Servir frontend ───────────────────────────────────────────────────────────

@app.route('/')
def index():
    index_file = FRONTEND_DIR / 'financas.html'
    if index_file.exists():
        return send_from_directory(str(FRONTEND_DIR), 'financas.html')
    return jsonify({
        'api': 'online',
        'message': 'Frontend não está neste deploy. Use os endpoints /api/*.'
    }), 200

@app.route('/<path:filename>')
def static_files(filename):
    static_file = FRONTEND_DIR / filename
    if static_file.exists():
        return send_from_directory(str(FRONTEND_DIR), filename)
    return jsonify({'erro': 'Arquivo não encontrado'}), 404


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("\n" + "="*55)
    print("  💰 CONTROLE FINANCEIRO — INICIANDO")
    print("="*55)

    if not DATABASE_URL:
        print("  Configure a variável de ambiente no Railway.")
    else:
        print("  ✅ Banco de dados conectado")

    print("  🚀 Rodando em http://0.0.0.0:5000")
    print("="*55 + "\n")

    app.run(host='0.0.0.0', port=5000, debug=False)
